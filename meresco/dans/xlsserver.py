## begin license ##
#
# "Meresco Components" are components to build searchengines, repositories
# and archives, based on "Meresco Core".
#
# Copyright (C) 2007-2009 SURF Foundation. http://www.surf.nl
# Copyright (C) 2007 SURFnet. http://www.surfnet.nl
# Copyright (C) 2007-2010 Seek You Too (CQ2) http://www.cq2.nl
# Copyright (C) 2007-2009 Stichting Kennisnet Ict op school. http://www.kennisnetictopschool.nl
# Copyright (C) 2012-2013, 2015, 2017 Seecr (Seek You Too B.V.) http://seecr.nl
# Copyright (C) 2012 Stichting Bibliotheek.nl (BNL) http://stichting.bibliotheek.nl
# Copyright (C) 2015 Stichting Kennisnet http://www.kennisnet.nl
# Copyright (C) 2017 SURF http://www.surf.nl
#
# This file is part of "Meresco Components"
#
# "Meresco Components" is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# "Meresco Components" is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with "Meresco Components"; if not, write to the Free Software
# Foundation, Inc., 51 Franklin St, Fifth Floor, Boston, MA  02110-1301  USA
#
## end license ##

from os.path import isfile, join, normpath, commonprefix, abspath, isdir
from rfc822 import formatdate
from time import time
from stat import ST_MTIME, ST_SIZE
from os import stat, listdir

from meresco.components.http import utils as httputils
from meresco.harvester.repositorystatus import RepositoryStatus
from meresco.components.http.utils import CRLF
from urllib import unquote, unquote_plus
from xml.sax.saxutils import quoteattr, escape as escapeHtml


from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
import os
import tempfile
import zipfile
from shutil import rmtree
from escaping import escapeFilename, unescapeFilename
import json


class File(object):
    def __init__(self, filename):
        self._filename = filename

    def exists(self):
        return isfile(self._filename)

    def getHeaders(self, expires=0):

        zipfilename = self._filename.split("/")[-1]
        contentType = 'application/octet-stream'

        return {
            'Date': self._date(),
            'Expires': self._date(expires),
            'Last-Modified': formatdate(stat(self._filename)[ST_MTIME]),
            'Content-Type': contentType,
            'Content-Disposition': 'attachment;filename="%s"' % zipfilename
        }

    def stream(self):
        yield 'HTTP/1.0 200 OK' + CRLF
        for item in self.getHeaders().items():
            yield "%s: %s" % item + CRLF
        yield CRLF

        fp = open(self._filename)
        data = fp.read(1024)
        while data:
            yield data
            data = fp.read(1024)
        fp.close()
        if isdir(os.path.dirname(self._filename)) and self._filename.startswith('/tmp/'):
            # print "Deleting tempdir..." + os.path.dirname(self._filename)
            rmtree(os.path.dirname(self._filename))

    def _date(self, offset=0):
        return formatdate(time() + offset)


class XlsServer(object):
    def __init__(self, name=None):
        self._name = name
        self._repostatus = RepositoryStatus('/var/log/meresco-harvester', '/var/lib/meresco-harvester/state')

    def handleRequest(self, path, port=None, Client=None, Method=None, Headers=None, **kwargs):
        resolvedFileOrDir = self._createXLS(kwargs.get("arguments")['rid'][0])
        if resolvedFileOrDir is None:
            yield httputils.notFoundHtml
            yield '<!DOCTYPE HTML PUBLIC "-//IETF//DTD HTML 2.0//EN">\n'
            yield "<html><head>\n"
            yield "<title>404 Not Found</title>\n"
            yield "</head><body>\n"
            yield "<h1>Not Found</h1>\n"
            yield "<p>The requested URL %s was not found on this server.</p>\n" % path
            yield "</body></html>\n"
            return
        yield resolvedFileOrDir.stream()


    def _createXLS(self, repositoryId):

        domainId = 'EduStandaard'
        invalidrecordIds = self._repostatus.invalidRecords(domainId, repositoryId)

        config = None
        repofilepath = self._getRepositoryJson(domainId, repositoryId)

        if repofilepath is not None and isfile(repofilepath):
            with open(repofilepath, 'r') as repojsonfile:
                config = json.load(repojsonfile)
            repojsonfile.close()
        
        if config is not None and invalidrecordIds:

            #Create workbook
            wb = Workbook()
            ws = wb.worksheets[0]

            ws.title = ('%s Invald OAI-PMH records' % repositoryId)[:30] #max. string length for worksheet title...

            # Add content:
            max_col = 0
            for idx, recId in enumerate(invalidrecordIds):
                recordId = recId.split(":", 1)[-1]
                if len(recordId) > max_col: max_col = len(recordId)
                etree = self._repostatus.getInvalidRecord(domainId, str(repositoryId), str(recordId))
                diagnostic = etree.xpath('//diag:diagnostic/diag:details/text()', namespaces={'diag':'http://www.loc.gov/zing/srw/diagnostic/'})                
                basisurl = "%(baseurl)s?verb=GetRecord&identifier=%(oai_id)s&metadataPrefix=%(mdpf)s" % {'baseurl': config['baseurl'], 'oai_id': recordId, 'mdpf': config['metadataPrefix']}
                firstcell = ws.cell(column=1, row=(idx+1), value=recordId)               
                firstcell.hyperlink = "%(baseurl)s?verb=GetRecord&identifier=%(oai_id)s&metadataPrefix=%(mdpf)s" % {'baseurl': config['baseurl'], 'oai_id': recordId, 'mdpf': config['metadataPrefix']}
                ws.cell(column=2, row=(idx+1), value=diagnostic[0])

            if max_col > 0: ws.column_dimensions['A'].width = max_col

            # Create tempdir to save to and archive to: 
            temp_dir = tempfile.mkdtemp()
            xlsPath = join(temp_dir, "%s.xlsx" % repositoryId)
            zipPath = join(temp_dir, "%s.zip" % repositoryId)

            wb.save(xlsPath)

            zf = zipfile.ZipFile(zipPath, mode='w')
            try:
              zf.write(xlsPath, "%s.xlsx" % repositoryId)
            finally:
              zf.close()

            return File(zipPath)

        return None

    def _getRepositoryJson(self, domainId, repositoryId):        
        repojsonfile = join('/var/lib/meresco-harvester/data', escapeFilename("%s.%s.repository" % (domainId, repositoryId)) )
        if not isfile(repojsonfile):
            return None
        return repojsonfile
